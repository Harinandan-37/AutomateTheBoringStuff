import openpyxl

filename = input("Excel File Directory: ")

wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

for row in range(1, sheet.max_row):
    for column in range(1, sheet.max_column):
        sheet_new.cell(row=column,column=row).value = sheet.cell(row=row,column=column).value


wb.save("Inverted.xlsx")
wb_new.save("Inverted.xlsx") #replaces file