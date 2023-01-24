import openpyxl

file_name = input("Enter file directory: ")
wb = openpyxl.load_workbook(file_name)
sheet = wb.active

num_files = sheet.max_column
num_char = sheet.max_row
file_list = []

for i in range(1,num_files+1):
    file_list.append("file"+str(i)+".txt")


print(file_list)

for row in range(1,num_files+1):
    for column in range(1,num_char+1):
        
        file_ = open(file_list[row-1],'a')
        file_.write(sheet.cell(row=column, column=row).value)
        print(sheet.cell(row=column, column=row).value,end='')

    print()
    

print("\nDone")

    

