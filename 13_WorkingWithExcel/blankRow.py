import openpyxl
import sys

filename = sys.argv[1]
row_num = int(sys.argv[2])
number = int(sys.argv[3])
    
wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb1 = openpyxl.Workbook()
new_sheet = wb1.active

for row in range(1, sheet.max_row+1):
    for column in range(1, sheet.max_column+1):
        if row <= row_num:
            new_sheet.cell(row=row,column=column).value = sheet.cell(row=row,column=column).value
        elif row > row_num:
            new_sheet.cell(row=row+number,column=column).value = sheet.cell(row=row,column=column).value


wb1.save('new_excel.xlsx')
wb.save('insertBlank.xlsx')

