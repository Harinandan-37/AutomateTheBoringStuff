import sys
import openpyxl
from openpyxl.styles import Font

N = int(sys.argv[1])
wb = openpyxl.Workbook()

sheet = wb.active




for row in range(1, N+2):
    for column in range(1, N+2):
        if column == 1 and row == 1:
            sheet.cell(row=1,column=1).value = ''
        else:
            sheet.cell(row=row,column=column).value = (row-1)*(column-1)

for fill_row in range(2,N+2):
    sheet.cell(row=1, column=fill_row).value = fill_row-1

for fill_column in range(2,N+2):
    sheet.cell(row=fill_column, column=1).value = fill_column-1

for row in range(1,N+2):
    for column in range(1,N+2):
        if row==1:
            sheet.cell(row=1, column=column).font = Font(name='Calibri', bold=True)
        elif column==1:
            sheet.cell(row=row, column=1).font = Font(name='Calibri', bold=True)

wb.save('MultiplicationTable.xlsx')