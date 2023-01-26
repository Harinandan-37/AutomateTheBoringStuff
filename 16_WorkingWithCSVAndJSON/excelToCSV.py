import openpyxl
import csv
import os


path = os.getcwd()

for files in os.listdir(path):
    if not files.endswith('.xlsx'):
        continue
    filename = os.path.basename(files)
    filename = filename.replace('.xlsx','')
    wb = openpyxl.load_workbook(files)

    for sheets in wb.sheetnames:

        csvName = filename + '_' + sheets + '.csv'
        csvFile = open(csvName, 'w', newline='')
        csvWriter = csv.writer(csvFile)
        sheet = wb.active

        for row in range(1,sheet.max_row+1):
            lst = []
            for column in range(1,sheet.max_column+1):
                lst.append(sheet.cell(row=row,column=column).value)
            csvWriter.writerow(lst)
        csvFile.close()